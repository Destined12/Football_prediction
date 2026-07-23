import os
import pandas as pd

RESULTS_DIR = os.path.join(os.path.dirname(__file__), '..', 'results')

FEATURE_DESCRIPTIONS = {
    'home_form_wins': ('Home team wins in last 5 matches', 'Count of wins for home team in last 5', 'int'),
    'home_form_draws': ('Home team draws in last 5', 'Count of draws for home team in last 5', 'int'),
    'home_form_losses': ('Home team losses in last 5', 'Count of losses for home team in last 5', 'int'),
    'home_form_points': ('Home team points in last 5', 'Wins*3 + Draws for home team in last 5', 'int'),
    'home_form_goals_scored': ('Home team goals scored last 5', 'Total goals scored by home team in last 5', 'int'),
    'home_form_goals_conceded': ('Home team goals conceded last 5', 'Total goals conceded by home team in last 5', 'int'),
    'home_form_goal_diff': ('Home team goal difference last 5', 'Goals scored - Goals conceded in last 5', 'int'),
    'home_form_avg_goals': ('Home team avg goals last 5', 'Goals scored / matches in last 5', 'float'),
    'home_form_avg_shots': ('Home team avg shots last 5', 'Total shots / matches in last 5', 'float'),
    'home_form_avg_sot': ('Home team avg shots on target last 5', 'Shots on target / matches in last 5', 'float'),
    'away_form_wins': ('Away team wins in last 5', 'Count of wins for away team in last 5', 'int'),
    'away_form_draws': ('Away team draws in last 5', 'Count of draws for away team in last 5', 'int'),
    'away_form_losses': ('Away team losses in last 5', 'Count of losses for away team in last 5', 'int'),
    'away_form_points': ('Away team points in last 5', 'Wins*3 + Draws for away team in last 5', 'int'),
    'away_form_goals_scored': ('Away team goals scored last 5', 'Total goals scored by away team in last 5', 'int'),
    'away_form_goals_conceded': ('Away team goals conceded last 5', 'Total goals conceded by away team in last 5', 'int'),
    'away_form_goal_diff': ('Away team goal difference last 5', 'Goals scored - Goals conceded in last 5', 'int'),
    'away_form_avg_goals': ('Away team avg goals last 5', 'Goals scored / matches in last 5', 'float'),
    'away_form_avg_shots': ('Away team avg shots last 5', 'Total shots / matches in last 5', 'float'),
    'away_form_avg_sot': ('Away team avg shots on target last 5', 'Shots on target / matches in last 5', 'float'),
    'home_team_home_win_rate': ('Home team historical win rate at home', 'Home wins / total home matches historically', 'float'),
    'home_team_home_avg_goals': ('Home team avg goals at home', 'Avg goals scored by home team in home matches', 'float'),
    'home_team_home_avg_goals_conceded': ('Home team avg goals conceded at home', 'Avg goals conceded by home team at home', 'float'),
    'home_team_home_avg_shots': ('Home team avg shots at home', 'Avg shots by home team in home matches', 'float'),
    'home_team_home_avg_sot': ('Home team avg shots on target at home', 'Avg SOT by home team in home matches', 'float'),
    'away_team_away_win_rate': ('Away team historical win rate away', 'Away wins / total away matches historically', 'float'),
    'away_team_away_avg_goals': ('Away team avg goals away', 'Avg goals scored by away team in away matches', 'float'),
    'away_team_away_avg_goals_conceded': ('Away team avg goals conceded away', 'Avg goals conceded by away team away', 'float'),
    'away_team_away_avg_shots': ('Away team avg shots away', 'Avg shots by away team in away matches', 'float'),
    'away_team_away_avg_sot': ('Away team avg shots on target away', 'Avg SOT by away team in away matches', 'float'),
    'h2h_home_wins': ('H2H home team wins (last 5)', 'Number of home team wins in last 5 meetings', 'int'),
    'h2h_away_wins': ('H2H away team wins (last 5)', 'Number of away team wins in last 5 meetings', 'int'),
    'h2h_draws': ('H2H draws (last 5)', 'Number of draws in last 5 meetings', 'int'),
    'h2h_home_goals': ('H2H home team goals (last 5)', 'Goals scored by home team in last 5 meetings', 'int'),
    'h2h_away_goals': ('H2H away team goals (last 5)', 'Goals scored by away team in last 5 meetings', 'int'),
    'h2h_goal_diff': ('H2H goal difference (last 5)', 'Home goals - away goals in last 5 meetings', 'int'),
    'home_atk_avg_goals': ('Home team attacking avg goals', 'Avg goals in last 10 matches (all venues)', 'float'),
    'home_atk_avg_shots': ('Home team attacking avg shots', 'Avg shots in last 10 matches', 'float'),
    'home_atk_avg_sot': ('Home team attacking avg SOT', 'Avg shots on target in last 10 matches', 'float'),
    'home_atk_avg_corners': ('Home team attacking avg corners', 'Avg corners in last 10 matches', 'float'),
    'away_atk_avg_goals': ('Away team attacking avg goals', 'Avg goals in last 10 matches (all venues)', 'float'),
    'away_atk_avg_shots': ('Away team attacking avg shots', 'Avg shots in last 10 matches', 'float'),
    'away_atk_avg_sot': ('Away team attacking avg SOT', 'Avg shots on target in last 10 matches', 'float'),
    'away_atk_avg_corners': ('Away team attacking avg corners', 'Avg corners in last 10 matches', 'float'),
    'home_def_avg_goals_conceded': ('Home team defensive avg goals conceded', 'Avg goals conceded in last 10 matches', 'float'),
    'home_def_clean_sheets': ('Home team clean sheet rate', 'Proportion of clean sheets in last 10', 'float'),
    'home_def_shots_faced': ('Home team avg shots faced', 'Avg shots faced in last 10 matches', 'float'),
    'home_def_avg_fouls': ('Home team avg fouls committed', 'Avg fouls in last 10 matches', 'float'),
    'home_def_avg_yellow': ('Home team avg yellow cards', 'Avg yellow cards in last 10 matches', 'float'),
    'home_def_avg_red': ('Home team avg red cards', 'Avg red cards in last 10 matches', 'float'),
    'away_def_avg_goals_conceded': ('Away team defensive avg goals conceded', 'Avg goals conceded in last 10 matches', 'float'),
    'away_def_clean_sheets': ('Away team clean sheet rate', 'Proportion of clean sheets in last 10', 'float'),
    'away_def_shots_faced': ('Away team avg shots faced', 'Avg shots faced in last 10 matches', 'float'),
    'away_def_avg_fouls': ('Away team avg fouls committed', 'Avg fouls in last 10 matches', 'float'),
    'away_def_avg_yellow': ('Away team avg yellow cards', 'Avg yellow cards in last 10 matches', 'float'),
    'away_def_avg_red': ('Away team avg red cards', 'Avg red cards in last 10 matches', 'float'),
    'home_disc_fouls': ('Home team discipline - fouls', 'Avg fouls committed in last 10 matches', 'float'),
    'home_disc_yellow': ('Home team discipline - yellows', 'Avg yellow cards in last 10 matches', 'float'),
    'home_disc_red': ('Home team discipline - reds', 'Avg red cards in last 10 matches', 'float'),
    'away_disc_fouls': ('Away team discipline - fouls', 'Avg fouls committed in last 10 matches', 'float'),
    'away_disc_yellow': ('Away team discipline - yellows', 'Avg yellow cards in last 10 matches', 'float'),
    'away_disc_red': ('Away team discipline - reds', 'Avg red cards in last 10 matches', 'float'),
    'home_mom_win_streak': ('Home team current win streak', 'Consecutive wins from last match backwards', 'int'),
    'home_mom_loss_streak': ('Home team current loss streak', 'Consecutive losses from last match backwards', 'int'),
    'home_mom_unbeaten_streak': ('Home team unbeaten streak', 'Consecutive unbeaten matches from last backwards', 'int'),
    'away_mom_win_streak': ('Away team current win streak', 'Consecutive wins from last match backwards', 'int'),
    'away_mom_loss_streak': ('Away team current loss streak', 'Consecutive losses from last match backwards', 'int'),
    'away_mom_unbeaten_streak': ('Away team unbeaten streak', 'Consecutive unbeaten matches from last backwards', 'int'),
}


def generate_feature_reports():
    eng_path = os.path.join(RESULTS_DIR, 'engineered_features', 'engineered_features.csv')
    df = pd.read_csv(eng_path)

    meta_cols = ['home_team', 'away_team', 'date', 'ftr', 'ftr_encoded', 'season']
    feature_cols = [c for c in df.columns if c not in meta_cols]

    desc_rows = []
    for col in sorted(feature_cols):
        if col in FEATURE_DESCRIPTIONS:
            name, desc, ftype = FEATURE_DESCRIPTIONS[col]
        else:
            name = col
            desc = f'Engineered feature: {col}'
            ftype = 'float'

        desc_rows.append({
            'Feature Name': col,
            'Description': desc,
            'Data Type': ftype,
            'Min': round(df[col].min(), 4) if col in df.columns else '',
            'Max': round(df[col].max(), 4) if col in df.columns else '',
            'Mean': round(df[col].mean(), 4) if col in df.columns else '',
            'Std': round(df[col].std(), 4) if col in df.columns else '',
        })

    desc_df = pd.DataFrame(desc_rows)
    desc_dir = os.path.join(RESULTS_DIR, 'feature_reports')
    os.makedirs(desc_dir, exist_ok=True)

    desc_df.to_csv(os.path.join(desc_dir, 'feature_descriptions.csv'), index=False)
    print(f'  Saved feature_descriptions.csv ({len(desc_df)} features)')

    stats_df = df[feature_cols].describe().T
    stats_df.to_csv(os.path.join(desc_dir, 'feature_statistics.csv'))
    print(f'  Saved feature_statistics.csv')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Feature Dictionary'

        headers = ['Feature Name', 'Description', 'Data Type', 'Formula', 'Purpose']
        header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for row_idx, (_, row) in enumerate(desc_df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row['Feature Name']).border = thin_border
            ws.cell(row=row_idx, column=2, value=row['Description']).border = thin_border
            ws.cell(row=row_idx, column=3, value=row['Data Type']).border = thin_border
            ws.cell(row=row_idx, column=4, value='Historical aggregation').border = thin_border
            ws.cell(row=row_idx, column=5, value='Match prediction').border = thin_border

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20

        wb.save(os.path.join(desc_dir, 'feature_dictionary.xlsx'))
        print(f'  Saved feature_dictionary.xlsx')
    except Exception as e:
        print(f'  Could not create Excel: {e}')

    print(f'\n  Feature reports generated successfully!')


if __name__ == '__main__':
    generate_feature_reports()
